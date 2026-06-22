import os
from xml.dom import minidom
from datetime import datetime
from openpyxl import load_workbook

class Produto:
    def __init__(self, tags):
        self.tags = tags
    
    def processar_itens(self, xml):
        itens = xml.getElementsByTagName("det")
        produtos = []
        
        for item in itens:
            produtos_dict = {}
            
            for tag in self.tags:
                elemento = item.getElementsByTagName(tag)
                if elemento and elemento[0].firstChild is not None:
                    nfe_content = elemento[0].firstChild.data
                    
                    # Formatação do NCM
                    if tag == "NCM":
                        ncm_formatado = nfe_content
                        nfe_content = f"{ncm_formatado[:4]}.{ncm_formatado[4:6]}.{ncm_formatado[6:8]}"
                    
                    # Formatação do valor unitário com 2 casas decimais
                    if tag == "vUnCom":
                        nfe_content = f"{float(nfe_content):.2f}"
                    
                    # Formatação do frete, seguro e outras despesas com 2 casas decimais
                    if tag in ["vFrete", "vSeg", "vOutro"]:
                        try:
                            nfe_content = f"{float(nfe_content):.2f}"
                        except ValueError:
                            nfe_content = "0.00" # Caso não seja um número válido
                    
                    # Formatação da origem sem zero à esquerda
                    if tag == "orig":
                        nfe_content = str(int(nfe_content))
                    
                    # Formatação do CST com zero à esquerda se necessário
                    if tag == "CST":
                        nfe_content = f"{int(nfe_content):02d}"
                    
                    # Formatação do percentual de ICMS
                    if tag == "pICMS":
                        try:
                            vlr_icms = float(nfe_content.replace(",", ".")) / 100
                            nfe_content = str(int(vlr_icms)) if vlr_icms.is_integer() else str(float(vlr_icms))
                        except ValueError:
                            pass

                    # Conversão de tipos para tags que não são strings especiais
                    if tag not in ("NCM", "orig", "CST"):
                        try:
                            produtos_dict[tag] = int(nfe_content)
                        except ValueError:
                            try:
                                produtos_dict[tag] = float(nfe_content)
                            except ValueError:
                                produtos_dict[tag] = nfe_content
                    else:
                        produtos_dict[tag] = nfe_content
                else:
                    # Se não encontrou o elemento, definir valor padrão
                    if tag == "vIPI":
                        produtos_dict[tag] = 0.00 # Definir como float para consistência
                    elif tag in ("NCM", "orig", "CST"):
                        produtos_dict[tag] = ""
                    elif tag in ["vFrete", "vSeg", "vOutro"]:
                        produtos_dict[tag] = 0.00 # Frete, seguro e outras despesas por item podem não existir
                    else:
                        produtos_dict[tag] = 0
            produtos.append(produtos_dict)
        return produtos

class Nfe:
    def __init__(self):
        # Tags específicas de cada produto/item, incluindo vFrete, vSeg, vOutro
        self.tags_produto = [
            "xProd","NCM", "CEST", "CFOP","qCom","vUnCom", "vDesc", "vProd", "orig", "CST", "pICMS", "vICMS","vBCST","pMVAST", "vICMSST", "vIPI", "vFrete", "vSeg", "vOutro"
        ]
        # Tags gerais da nota (sem os valores de rateio)
        self.tags_notas = [
            "chNFe", "nNF", "dhEmi", "UF","CRT"
        ]
        # Ordem correta das colunas na planilha
        self.tags_ordem_certa = [
            "chNFe","nNF","dhEmi","UF","CRT",
            "xProd","NCM","CEST","CFOP","qCom","vUnCom","vFrete","vSeg","vDesc","vOutro","vProd","vIPI",
            "orig","CST","pICMS","vICMS","vBCST","vICMSST","pMVAST"
        ]
        
    def formatarNfe(self, arquivos):
        lista_formatada = []
        for arquivo in arquivos:
            with open(arquivo, "r", encoding="utf-8") as nfe_files:
                xml = minidom.parse(nfe_files)
                
                # Processar dados gerais da nota
                dados_notas = {}
                for tag in self.tags_notas:
                    elemento = xml.getElementsByTagName(tag)
                    if elemento and elemento[0].firstChild is not None:
                        valor = elemento[0].firstChild.data

                        if tag == "dhEmi":
                            data = datetime.fromisoformat(valor)
                            valor = f"{data.day}/{data.month}/{data.year}"

                        if tag == "CRT":
                            if valor in ["1", "2"]:
                                valor = "Simples Nacional"
                            elif valor == "3":
                                valor = "Regime Normal"
                        
                        if tag in ["nNF"]:
                            try:
                                valor = int(valor.replace(",", "."))
                            except ValueError:
                                try:
                                    valor = float(valor.replace(",", "."))
                                except Exception:
                                    pass
                        dados_notas[tag] = valor

                # Processar produtos/itens
                produto_processados = Produto(self.tags_produto)
                produtos = produto_processados.processar_itens(xml)
                
                # Combinar dados da nota com dados de cada produto
                nfe_ordenada = []
                for p in produtos:
                    combinar = {**dados_notas, **p}
                    dicionario_ordenado = {tag: combinar.get(tag, 0) for tag in self.tags_ordem_certa}
                    nfe_ordenada.append(dicionario_ordenado)

                lista_formatada.append(nfe_ordenada)
        return lista_formatada

class NFEProcessor:
    def __init__(self):
        self.alfabeto = []

    def gerarAlfabeto(self):
        self.alfabeto = []
        for k in range(ord("A"), ord("Z")+1):
            self.alfabeto.append(chr(k))
        for k1 in range(ord("A"), ord("Z")+1):
            for k2 in range(ord("A"), ord("Z")+1):
                coluna = chr(k1) + chr(k2)
                self.alfabeto.append(coluna)
                if coluna == "BG":
                    break
            if coluna == "BG":
                break
        return self.alfabeto

    def _extrair_cnpj_e_competencia(self, primeiro_xml_path):
        cnpj = None
        competencia = None
        try:
            with open(primeiro_xml_path, "r", encoding="utf-8") as f:
                xml = minidom.parse(f)
                
                # Buscar CNPJ dentro da tag <dest>
                dest_nodes = xml.getElementsByTagName("dest")
                if dest_nodes:
                    cnpj_nodes = dest_nodes[0].getElementsByTagName("CNPJ")
                    if cnpj_nodes and cnpj_nodes[0].firstChild is not None:
                        cnpj_raw = cnpj_nodes[0].firstChild.data
                        cnpj = "".join([c for c in cnpj_raw if c.isdigit()])

                # Competência (mês-ano a partir do dhEmi)
                dh_nodes = xml.getElementsByTagName("dhEmi")
                if dh_nodes and dh_nodes[0].firstChild is not None:
                    dh_text = dh_nodes[0].firstChild.data
                    dt = datetime.fromisoformat(dh_text)
                    competencia = f"{dt.month:02}-{dt.year}"
        except Exception:
            pass
        return cnpj, competencia

    def process_files(self, arquivos_xml):
        nfe = Nfe()
        lista_formatada = nfe.formatarNfe(arquivos_xml)
        alfabeto = self.gerarAlfabeto()

        # Carregar a planilha base
        template_path = os.path.join(os.path.dirname(__file__), "planilha_mae/teste.xlsx")
        wb = load_workbook(template_path)
        sheet_lerxml = wb["LerXml"]
        sheet_relatorio = wb["Relatorio"] if "Relatorio" in wb.sheetnames else wb.create_sheet("Relatorio")

        # Extrair CNPJ e competência do primeiro XML
        cnpj, competencia = self._extrair_cnpj_e_competencia(arquivos_xml[0])
        if not cnpj:
            cnpj = "00000000000000"
        if not competencia:
            now = datetime.now()
            competencia = f"{now.month:02}-{now.year}"

        # Escrever CNPJ na célula A1 de todas as abas
        for ws in wb.worksheets:
            ws["A1"] = cnpj

        # Inserir dados a partir da linha 3 na aba LerXml
        j = 3
        for arquivo in lista_formatada:
            for produto in arquivo:
                i_col = 0
                for chave, valor in produto.items():
                    sheet_lerxml[f"{alfabeto[i_col]}{j}"] = valor
                    i_col += 1
                # Adicionar 'REVENDA' na coluna Y (índice 24, pois Y é a 25ª letra)
                sheet_lerxml[f"{alfabeto[24]}{j}"] = "REVENDA"
                j += 1
            j += 1
            # Não pular linha entre arquivos na aba LerXml

        # Para a aba Relatorio: apenas um item por nota (o primeiro item de cada nota)
        j_relatorio = 3
        for arquivo in lista_formatada:
            if arquivo:  # Se há itens neste arquivo
                primeiro_item = arquivo[0]  # Pegar apenas o primeiro item
                # Inserir apenas chNFe e nNF na aba Relatorio
                sheet_relatorio[f"A{j_relatorio}"] = primeiro_item.get("chNFe", "")
                sheet_relatorio[f"B{j_relatorio}"] = primeiro_item.get("nNF", "")
                sheet_relatorio[f"C{j_relatorio}"] = primeiro_item.get("dhEmi", "")
                sheet_relatorio[f"Y{j_relatorio}"] = "REVENDA"
                j_relatorio += 1

        # Nome do arquivo: "cnpj competencia - dia-mes-ano hora-minuto.xlsx"
        now = datetime.now()
        dia_hora = f"{now.day:02}-{now.month:02}-{now.year} {now.hour:02}-{now.minute:02}"
        filename = f"{cnpj} {competencia} - {dia_hora}.xlsx"

        downloads_dir = os.path.join(os.path.dirname(__file__), "downloads")
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        filepath = os.path.join(downloads_dir, filename)

        wb.save(filepath)
        return filepath
